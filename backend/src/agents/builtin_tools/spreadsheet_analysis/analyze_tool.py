"""Analyze spreadsheet files using Code Interpreter.

Factory function creates a context-bound tool that downloads tabular files
from S3, pushes them to Code Interpreter, and executes Python code for analysis.
"""

import logging
import os
import re
from typing import Any, Dict, Optional

import boto3
from strands import tool

from .list_spreadsheets_tool import _get_kb_files, _get_session_files

logger = logging.getLogger(__name__)

MAX_OUTPUT_CHARS = 10000  # ~2500 tokens — safe margin under context limits
MAX_ERROR_CHARS = 600  # cleaned traceback budget — full pandas tracebacks are noise

# Defensive caps for multi-sheet XLSX conversion. The outer upload limit
# (FILE_UPLOAD_MAX_SIZE_BYTES, default 4 MB) catches naive abuse, but XLSX
# is a zip of XML and can pack thousands of nearly-empty sheets into a few
# megabytes. We cap both sheet count and per-sheet row count to keep turn
# latency bounded; anything excluded is surfaced to the model with a
# warning so the user learns about the cap rather than getting silently
# wrong results.
MAX_SHEETS_TO_CONVERT = int(os.environ.get("ANALYZE_MAX_SHEETS", 25))
MAX_ROWS_PER_SHEET = int(os.environ.get("ANALYZE_MAX_ROWS_PER_SHEET", 500_000))

_SCHEMA_MARKER = "[__SCHEMA__]"
_SHEETS_MARKER = "[__SHEETS__]"


def _sanitize_sheet_name(name: str) -> str:
    """Make a sheet name filesystem-safe.

    Sheet names can contain spaces, slashes, unicode — pick a deterministic
    filename-safe transform so the model can predict the output filename
    from the sheet name. Lowercase for cross-platform stability, replace
    anything non-alphanumeric with underscore, collapse repeats, trim.
    """
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").lower()
    return cleaned or "sheet"


def _parse_sheet_inventory(bootstrap_stdout: str) -> Dict[str, Any]:
    """Extract the sheet inventory emitted by the XLSX bootstrap.

    The bootstrap prints a pipe-delimited record per converted sheet
    (``sheet|<name>|<path>|<rows>|<truncated_flag>|<primary_alias>``),
    bracketed by ``_SHEETS_MARKER``. We parse that into a structured dict
    the tool can reason about without re-evaluating Python literals from
    untrusted-ish interpreter stdout.

    Returns a dict with:
        - ``total`` (int): total sheets in workbook
        - ``converted`` (int): sheets actually written to CSV
        - ``skipped`` (int): sheets excluded by MAX_SHEETS_TO_CONVERT
        - ``skipped_preview`` (list[str]): first few skipped sheet names
        - ``sheets`` (list[dict]): per-sheet records with name, path,
          rows, truncated
        - ``has_primary_alias`` (bool): whether the <stem>.csv fast-path
          alias was written for the first sheet
    """
    result: Dict[str, Any] = {
        "total": 0,
        "converted": 0,
        "skipped": 0,
        "skipped_preview": [],
        "sheets": [],
        "has_primary_alias": False,
    }
    if _SHEETS_MARKER not in bootstrap_stdout:
        return result
    try:
        block = bootstrap_stdout.split(_SHEETS_MARKER)[1].strip()
    except IndexError:
        return result

    for line in block.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("total:"):
            result["total"] = _safe_int(stripped.split(":", 1)[1])
        elif stripped.startswith("converted:"):
            result["converted"] = _safe_int(stripped.split(":", 1)[1])
        elif stripped.startswith("skipped:"):
            result["skipped"] = _safe_int(stripped.split(":", 1)[1])
        elif stripped.startswith("skipped_names:"):
            # Stored as a Python list literal — safe to ast.literal_eval
            # because the content is quoted strings from sheetnames.
            import ast as _ast
            try:
                names = _ast.literal_eval(stripped.split(":", 1)[1].strip())
                if isinstance(names, list):
                    result["skipped_preview"] = [str(n) for n in names]
            except (ValueError, SyntaxError):
                pass
        elif stripped.startswith("sheet|"):
            parts = stripped.split("|")
            # sheet | name | path | rows | truncated | alias
            if len(parts) < 6:
                continue
            _, name, path, rows, trunc, alias = parts[:6]
            result["sheets"].append({
                "name": name,
                "path": path,
                "rows": _safe_int(rows),
                "truncated": trunc == "1",
                "primary_alias": alias or None,
            })
            if alias:
                result["has_primary_alias"] = True
    return result


def _safe_int(raw: str) -> int:
    try:
        return int(str(raw).strip())
    except (TypeError, ValueError):
        return 0


def _format_sheet_note(inventory: Dict[str, Any]) -> str:
    """Turn a parsed sheet inventory into a markdown footer for the tool
    response. Empty string when the workbook has a single sheet that was
    fully converted (no-op path — nothing interesting to report).
    """
    total = inventory.get("total", 0)
    sheets = inventory.get("sheets", [])
    skipped = inventory.get("skipped", 0)
    truncated_sheets = [s for s in sheets if s.get("truncated")]

    if total <= 1 and not truncated_sheets:
        return ""

    lines: list[str] = []

    if total > 1:
        converted = inventory.get("converted", len(sheets))
        if skipped:
            preview = inventory.get("skipped_preview", [])
            shown = ", ".join(preview)
            more = f" (+{skipped - len(preview)} more)" if skipped > len(preview) else ""
            lines.append(
                f"⚠ Workbook has {total} sheets; converted the first {converted}. "
                f"Skipped: {shown}{more}. "
                f"Split the file or export specific tabs as CSV to analyze the rest."
            )
        else:
            lines.append(
                f"Workbook has {total} sheets; all converted. Use the "
                f"per-sheet filenames below to read or combine them."
            )
        lines.append("")
        lines.append("**Available sheets (load via `pd.read_csv`):**")
        for s in sheets:
            trunc_tag = ""
            if s.get("truncated"):
                trunc_tag = f" — ⚠ truncated at {MAX_ROWS_PER_SHEET:,} rows"
            lines.append(f"- `{s['name']}` → `{s['path']}` ({s['rows']:,} rows{trunc_tag})")

    elif truncated_sheets:
        # Single-sheet workbook but hit the row cap.
        s = truncated_sheets[0]
        lines.append(
            f"⚠ Sheet `{s['name']}` was truncated at {MAX_ROWS_PER_SHEET:,} rows "
            f"due to the analysis size cap; full row count not reported."
        )

    return "\n".join(lines)


def _truncate_output(text: str) -> str:
    """Truncate tool output to prevent blowing the LLM context window."""
    if not text or len(text) <= MAX_OUTPUT_CHARS:
        return text
    return text[:MAX_OUTPUT_CHARS] + f"\n\n... (output truncated — {len(text):,} chars total, showing first {MAX_OUTPUT_CHARS:,})"


def _strip_first_row(schema: str) -> str:
    """Drop the ``first_row: ...`` line from a schema footer.

    On the happy path the first-row preview helps the model write correct
    code. On the error path the model already has the load line and column
    list — the full row dump is ~30 fields of noise. This trims it.
    """
    return "\n".join(
        line for line in schema.splitlines()
        if not line.startswith("first_row:")
    )


# ---------------------------------------------------------------------------
# Stderr cleaning
# ---------------------------------------------------------------------------

# Frames we never want to show the LLM — they're pandas/numpy internals with
# zero signal for fixing the user's code.
_INTERNAL_FRAME_MARKERS = (
    "site-packages/pandas/",
    "site-packages/numpy/",
    "pandas/_libs/",
    "pandas/core/",
    "pandas/io/",
)


def _clean_stderr(stderr: str) -> str:
    """Strip pandas internal frames and dtype warnings from a traceback.

    Keeps the user-code frame (the `/tmp/ipykernel_*.py` line they wrote) and
    the final exception line. Falls back to a truncated raw stderr if the
    traceback doesn't match the expected shape.
    """
    if not stderr:
        return "Unknown error"

    lines = stderr.splitlines()

    # 1. Drop DtypeWarning noise (spans 2 lines: the warning + the call-site).
    filtered: list[str] = []
    skip_next = False
    for line in lines:
        if skip_next:
            skip_next = False
            continue
        if "DtypeWarning:" in line or "FutureWarning:" in line or "UserWarning:" in line:
            skip_next = True  # next line is usually the offending code snippet
            continue
        filtered.append(line)

    # 2. Find the final exception line (e.g. "KeyError: 'NET_AMOUNT'").
    final_exception = ""
    for line in reversed(filtered):
        stripped = line.strip()
        if not stripped:
            continue
        # Exception lines are left-flush and match "ExceptionName: message".
        if not line.startswith((" ", "\t")) and re.match(r"^[A-Z][A-Za-z]*(?:Error|Exception|Warning):", stripped):
            final_exception = stripped
            break

    # 3. Find the user-code frame (ipykernel tempfile, not site-packages).
    user_frame_lines: list[str] = []
    for i, line in enumerate(filtered):
        stripped = line.strip()
        if not stripped.startswith("File "):
            continue
        if any(m in stripped for m in _INTERNAL_FRAME_MARKERS):
            continue
        # Keep this frame + up to the next 2 lines (the code snippet + pointer).
        user_frame_lines.append(stripped)
        for j in range(i + 1, min(i + 3, len(filtered))):
            nxt = filtered[j].strip()
            if not nxt or nxt.startswith("File "):
                break
            user_frame_lines.append(nxt)
        break

    if user_frame_lines and final_exception:
        cleaned = "\n".join(user_frame_lines) + "\n" + final_exception
    elif final_exception:
        cleaned = final_exception
    else:
        # Unrecognized shape — return a short tail rather than a 3K dump.
        cleaned = "\n".join(filtered[-8:]).strip()

    if len(cleaned) > MAX_ERROR_CHARS:
        # Leave room for the ellipsis tail so the final string respects
        # the budget strictly — callers rely on ``len(output) <=
        # MAX_ERROR_CHARS``.
        ellipsis = " ..."
        cleaned = cleaned[:MAX_ERROR_CHARS - len(ellipsis)] + ellipsis
    return cleaned


# ---------------------------------------------------------------------------
# Schema preview probe
# ---------------------------------------------------------------------------


def _build_preview_code(csv_filename: str) -> str:
    """Return Python code that prints a compact schema snapshot for csv_filename.

    Runs a bounded skiprows probe (0..8) to handle report-style exports with
    leading metadata rows. Picks the skiprows value that produces the cleanest
    header — no ``Unnamed:`` columns, no duplicates, non-empty names — and
    emits a ready-to-use ``pd.read_csv(...)`` invocation when the best
    candidate is meaningfully better than skiprows=0. Otherwise it reports the
    columns at skiprows=0 and lets the model decide.

    Output is bracketed with _SCHEMA_MARKER so it can be reliably extracted
    from the interpreter's stdout stream even if user code prints other things.

    Filenames with quotes or other f-string-breaking characters are handled
    by stashing the filename as a top-of-script local variable (``_FNAME``)
    via ``repr()`` once. The rest of the template references ``_FNAME`` as
    an ordinary string, so we never re-interpolate the raw filename into
    nested f-string contexts. Before this indirection, a filename like
    ``"O'Brien data.csv"`` would generate invalid Python because ``repr``
    emits double quotes when the string contains a single quote, conflicting
    with the outer f-string's own quoting.
    """
    # repr() always produces a valid Python string literal; storing that
    # literal once means the generated code can refer to the filename by
    # name, without any further escaping.
    fname_literal = repr(csv_filename)
    return f"""
import warnings, pandas as pd
warnings.filterwarnings('ignore')

_FNAME = {fname_literal}

def _score(cols):
    # Higher is better. Punishes Unnamed columns and duplicates.
    if not cols:
        return -10_000
    unnamed = sum(1 for c in cols if str(c).startswith('Unnamed:'))
    named = len(cols) - unnamed
    dup_penalty = (len(cols) - len(set(cols))) * 20
    blank_penalty = sum(1 for c in cols if not str(c).strip()) * 10
    return named - (unnamed * 5) - dup_penalty - blank_penalty

try:
    with open(_FNAME, 'r') as _fh:
        _total_rows = sum(1 for _ in _fh)

    # Score skiprows=0..8, keep the winner and remember the baseline.
    _baseline_score, _baseline_cols = -float('inf'), []
    _best_skip, _best_score, _best_cols = 0, -float('inf'), []
    for _sk in range(9):
        try:
            _cols = pd.read_csv(_FNAME, nrows=0, skiprows=_sk, low_memory=False).columns.tolist()
        except Exception:
            continue
        _sc = _score(_cols)
        if _sk == 0:
            _baseline_score, _baseline_cols = _sc, _cols
        if _sc > _best_score:
            _best_skip, _best_score, _best_cols = _sk, _sc, _cols

    # Confidence gate: only prescribe a non-zero skiprows when the winner
    # actually fixes a header problem — either more named columns OR fewer
    # Unnamed columns than the baseline — AND the winner is mostly clean.
    # A score-delta threshold alone can't distinguish "found the real header"
    # from "data row happens to parse cleanly", so we anchor on named/unnamed
    # counts instead.
    def _named_unnamed(cols):
        u = sum(1 for c in cols if str(c).startswith('Unnamed:'))
        return len(cols) - u, u
    _base_named, _base_unnamed = _named_unnamed(_baseline_cols)
    _win_named, _win_unnamed = _named_unnamed(_best_cols)
    _win_clean_ratio = _win_named / max(len(_best_cols), 1)

    _prescribe = (
        _best_skip > 0
        and _win_clean_ratio >= 0.7
        and (_win_named > _base_named or _win_unnamed < _base_unnamed)
    )

    if _prescribe:
        _report_skip, _report_cols = _best_skip, _best_cols
    else:
        _report_skip, _report_cols = 0, _baseline_cols

    _data_rows = max(_total_rows - 1 - _report_skip, 0)
    _col_preview = ', '.join(str(c) for c in _report_cols[:20])
    if len(_report_cols) > 20:
        _col_preview += f' ... (+{{len(_report_cols) - 20}} more)'

    try:
        _head = pd.read_csv(_FNAME, skiprows=_report_skip, nrows=1, low_memory=False)
        _first_row = _head.iloc[0].to_dict() if len(_head) else {{}}
        _first_row = {{k: (str(v)[:40] + '...' if len(str(v)) > 40 else v) for k, v in _first_row.items()}}
    except Exception:
        _first_row = {{}}

    if _prescribe:
        _load = f"pd.read_csv({{_FNAME!r}}, skiprows={{_report_skip}}, low_memory=False)"
        _note = f"  # {{_report_skip}} metadata row(s) detected before header"
    else:
        _load = f"pd.read_csv({{_FNAME!r}}, low_memory=False)"
        _note = ""

    print({_SCHEMA_MARKER!r})
    print(f'file: {{_FNAME}} ({{_data_rows}} rows x {{len(_report_cols)}} cols)')
    print(f'load: {{_load}}{{_note}}')
    print(f'columns: {{_col_preview}}')
    print(f'first_row: {{_first_row}}')
    # If confidence was low, flag it so the model knows to verify.
    if not _prescribe and _win_unnamed > 0 and _win_unnamed < len(_best_cols):
        print(f'note: header may need adjustment (skiprows=0 has {{_base_unnamed}}/{{len(_baseline_cols)}} unnamed columns); inspect head() if unsure')
    print({_SCHEMA_MARKER!r})
except Exception as _e:
    print({_SCHEMA_MARKER!r})
    print(f'schema preview unavailable: {{_e}}')
    print({_SCHEMA_MARKER!r})
"""


def _extract_schema_preview(stdout: str) -> tuple[str, str]:
    """Split stdout into (schema_block, remaining_stdout).

    The schema block is whatever is between _SCHEMA_MARKER pairs; if no markers
    are found, returns ("", stdout).
    """
    if _SCHEMA_MARKER not in stdout:
        return "", stdout
    parts = stdout.split(_SCHEMA_MARKER)
    # parts = [before, schema, after, ...]; stitch back everything non-schema.
    if len(parts) >= 3:
        schema = parts[1].strip()
        remaining = (parts[0] + _SCHEMA_MARKER.join(parts[2:])).strip("\n")
        return schema, remaining
    return "", stdout


def _get_code_interpreter_id() -> Optional[str]:
    """Get Code Interpreter ID from environment or SSM."""
    ci_id = os.getenv("AGENTCORE_CODE_INTERPRETER_ID")
    if ci_id:
        return ci_id
    try:
        project_name = os.getenv("PROJECT_NAME", "strands-agent-chatbot")
        environment = os.getenv("ENVIRONMENT", "dev")
        region = os.getenv("AWS_REGION", "us-west-2")
        ssm = boto3.client("ssm", region_name=region)
        response = ssm.get_parameter(Name=f"/{project_name}/{environment}/agentcore/code-interpreter-id")
        return response["Parameter"]["Value"]
    except Exception:
        return None


def make_analyze_tool(
    assistant_id: Optional[str],
    session_id: str,
    user_id: str,
):
    """Create an analyze_spreadsheet tool bound to the given context."""

    @tool
    async def analyze_spreadsheet(
        filename: str,
        python_code: str,
        output_filename: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Analyze a spreadsheet file using Python code in Code Interpreter.

        Downloads the specified file and loads it into a sandboxed Python
        environment for analysis. Use pandas, numpy, matplotlib, and seaborn.

        ⚠️  CRITICAL — filename vs. in-sandbox path
        -------------------------------------------
        The ``filename`` parameter names the **source** file (exactly as it
        appears in the chat attachment or knowledge base, e.g.
        ``"FY_27_Ledger.xlsx"``).

        In the sandbox, XLSX files are pre-converted to CSV:

        • Single-sheet workbooks → loadable as ``<stem>.csv``
          ``FY_27_Ledger.xlsx`` → ``FY_27_Ledger.csv``

        • Multi-sheet workbooks → one CSV per sheet, plus a primary alias
          for the first sheet:
            ``Budget.xlsx`` → ``Budget.summary.csv``,
                              ``Budget.transactions.csv``,
                              ``Budget.notes.csv``,
                              ``Budget.csv`` (alias of the first sheet)

          The tool response's "Available sheets" footer lists the exact
          ``pd.read_csv`` target for every sheet. **Use those names
          verbatim.** For cross-sheet aggregation, read multiple and
          combine with ``pd.concat``.

        So ``python_code`` must read the CSV form, even for an XLSX source:

            filename:    "FY_27_Ledger.xlsx"      (source name)
            python_code: pd.read_csv('FY_27_Ledger.csv', low_memory=False)
                                         ^^^ .csv, not .xlsx

        CSV files keep their name unchanged in the sandbox.

        Handling leading metadata rows
        ------------------------------
        Some exports have metadata rows above the real header. The tool
        response always includes a schema footer with a ready-to-use
        ``load:`` command that accounts for this — e.g.
        ``pd.read_csv('file.csv', skiprows=3, low_memory=False)``.
        **On any retry, use that exact load line verbatim** instead of
        guessing ``skiprows``.

        Safety limits
        -------------
        Multi-sheet workbooks convert at most the first 25 sheets; each
        sheet is truncated at 500,000 rows. When a cap triggers, the
        response footer tells you what was excluded so you can relay it
        to the user instead of presenting a partial answer as complete.

        Best for: aggregations, filtering, trends, comparisons, statistics,
        charts. For simple factual lookups, use knowledge base search.

        Args:
            filename: Source filename from list_spreadsheets results. Use
                the original name (``.xlsx`` or ``.csv``), not the sandbox
                form.
            python_code: Python to execute. For XLSX sources, use the exact
                CSV names from the "Available sheets" footer. Available
                libraries: pandas, numpy, matplotlib, seaborn, openpyxl.
            output_filename: Optional PNG filename if generating a chart.
                Must end with ``.png``. Example: ``"chart.png"``.

        Returns:
            Analysis results as text (with a schema footer), and optionally
            a chart image.
        """
        from bedrock_agentcore.tools.code_interpreter_client import CodeInterpreter

        # 1. Validate Code Interpreter is available
        ci_id = _get_code_interpreter_id()
        if not ci_id:
            return {"content": [{"text": "❌ Code Interpreter is not configured. Contact your administrator."}], "status": "error"}

        # 2. Find the file in accessible sources
        file_info = await _find_file(filename, assistant_id, session_id)
        if not file_info:
            return {"content": [{"text": f"❌ File '{filename}' not found or not accessible. Use list_spreadsheets to see available files."}], "status": "error"}

        # 3. Download from S3
        try:
            file_bytes = _download_file(file_info)
        except Exception as e:
            return {"content": [{"text": f"❌ Failed to download file: {e}"}], "status": "error"}

        # 4. Push file to Code Interpreter
        content_type = file_info.get("content_type", "")
        is_xlsx = "spreadsheetml" in content_type or filename.lower().endswith(".xlsx")

        region = os.getenv("AWS_REGION", "us-west-2")
        code_interpreter = CodeInterpreter(region)

        try:
            code_interpreter.start(identifier=ci_id)

            if is_xlsx:
                # Push XLSX as base64, decode in sandbox, then convert every
                # sheet to its own CSV (subject to defensive caps below).
                # Model gets a full sheet inventory in the schema footer so
                # cross-sheet aggregation works in a single analyze call.
                import base64
                b64_content = base64.b64encode(file_bytes).decode("ascii")
                stem = os.path.splitext(filename)[0]
                # Back-compat alias: single-sheet workbooks still expose
                # <stem>.csv so the one-file, one-sheet fast path keeps
                # its existing filename contract. Multi-sheet workbooks
                # use <stem>.<sanitized_sheet>.csv per sheet.
                primary_csv_filename = f"{stem}.csv"

                code_interpreter.invoke("writeFiles", {"content": [
                    {"path": "_encoded.b64", "text": b64_content},
                ]})
                # Bootstrap: iterate every sheet (capped), write a CSV per
                # sheet, emit an inventory the outer tool can parse. Uses
                # read_only + values_only to avoid loading full styles/
                # formulas into memory — important for large workbooks.
                bootstrap_code = f"""
import base64, io, csv, re
from openpyxl import load_workbook

MAX_SHEETS = {MAX_SHEETS_TO_CONVERT}
MAX_ROWS = {MAX_ROWS_PER_SHEET}
STEM = {stem!r}
PRIMARY_CSV = {primary_csv_filename!r}

def _sanitize(name):
    cleaned = re.sub(r'[^A-Za-z0-9]+', '_', name).strip('_').lower()
    return cleaned or 'sheet'

with open('_encoded.b64', 'r') as f:
    raw = base64.b64decode(f.read())

wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
all_sheets = wb.sheetnames
total_sheets = len(all_sheets)
sheets_to_convert = all_sheets[:MAX_SHEETS]
skipped_sheets = all_sheets[MAX_SHEETS:]

# Track which sanitized names we've used — de-duplicate if two sheet
# names sanitize to the same token (e.g. "Q1 2026" and "q1_2026").
used_names = set()
def _unique(base):
    candidate, n = base, 2
    while candidate in used_names:
        candidate = f"{{base}}_{{n}}"
        n += 1
    used_names.add(candidate)
    return candidate

# Single-sheet workbook: keep the legacy <stem>.csv filename for
# back-compat with existing prompts/docstring examples. Multi-sheet
# workbooks get <stem>.<sheet>.csv per sheet. The primary alias is
# always the first sheet.
sheet_records = []
for idx, sheet_name in enumerate(sheets_to_convert):
    if total_sheets == 1:
        out_path = PRIMARY_CSV
    else:
        safe = _unique(_sanitize(sheet_name))
        out_path = f"{{STEM}}.{{safe}}.csv"

    ws = wb[sheet_name]
    rows_written = 0
    truncated = False
    with open(out_path, 'w', newline='') as out:
        writer = csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            if rows_written >= MAX_ROWS:
                truncated = True
                break
            writer.writerow([str(cell) if cell is not None else '' for cell in row])
            rows_written += 1

    # Alias the first sheet of a multi-sheet workbook to the legacy
    # <stem>.csv path too, so the single-sheet fast path and the
    # XLSX→CSV docstring example keep working for picking "the main
    # sheet" without needing to know its name.
    primary_alias = None
    if total_sheets > 1 and idx == 0:
        try:
            with open(out_path, 'r') as src, open(PRIMARY_CSV, 'w') as dst:
                dst.write(src.read())
            primary_alias = PRIMARY_CSV
        except Exception:
            pass

    sheet_records.append({{
        'name': sheet_name,
        'path': out_path,
        'rows': rows_written,
        'truncated': truncated,
        'primary_alias': primary_alias,
    }})

print({_SHEETS_MARKER!r})
print(f'total: {{total_sheets}}')
print(f'converted: {{len(sheet_records)}}')
print(f'skipped: {{len(skipped_sheets)}}')
if skipped_sheets:
    _preview = skipped_sheets[:5]
    print(f'skipped_names: {{_preview}}')
for rec in sheet_records:
    # Emit one record per line, pipe-delimited, so the outer parser
    # doesn't have to evaluate arbitrary Python literals.
    trunc = '1' if rec['truncated'] else '0'
    alias = rec['primary_alias'] or ''
    print(f"sheet|{{rec['name']}}|{{rec['path']}}|{{rec['rows']}}|{{trunc}}|{{alias}}")
print({_SHEETS_MARKER!r})
wb.close()
"""
                resp = code_interpreter.invoke("executeCode", {"code": bootstrap_code, "language": "python", "clearContext": False})
                bootstrap_stdout = ""
                for event in resp.get("stream", []):
                    result = event.get("result", {})
                    if result.get("isError", False):
                        error_msg = _clean_stderr(result.get("structuredContent", {}).get("stderr", ""))
                        return {"content": [{"text": f"❌ Failed to convert XLSX in sandbox:\n```\n{error_msg}\n```"}], "status": "error"}
                    bootstrap_stdout += result.get("structuredContent", {}).get("stdout", "")

                sheet_inventory = _parse_sheet_inventory(bootstrap_stdout)
                if not sheet_inventory["sheets"]:
                    return {
                        "content": [{"text": "❌ XLSX bootstrap produced no readable sheets."}],
                        "status": "error",
                    }

                # csv_filename is the canonical name the rest of the code
                # path uses to probe schema and emit "load:" hints. For
                # single-sheet or the primary alias on multi-sheet, that's
                # <stem>.csv. For multi-sheet with no primary alias (write
                # failure), fall back to the first converted sheet's path.
                csv_filename = (
                    primary_csv_filename
                    if sheet_inventory["has_primary_alias"] or len(sheet_inventory["sheets"]) == 1
                    else sheet_inventory["sheets"][0]["path"]
                )
                multi_sheet_note = _format_sheet_note(sheet_inventory)
            else:
                # CSV — push directly as text
                csv_filename = filename if filename.lower().endswith(".csv") else os.path.splitext(filename)[0] + ".csv"
                multi_sheet_note = ""
                try:
                    csv_text = file_bytes.decode("utf-8")
                except UnicodeDecodeError:
                    csv_text = file_bytes.decode("utf-8", errors="replace")
                code_interpreter.invoke("writeFiles", {"content": [{"path": csv_filename, "text": csv_text}]})

            # 5. Probe schema — separate exec so its output is isolated from user code.
            schema_preview = ""
            try:
                preview_resp = code_interpreter.invoke("executeCode", {
                    "code": _build_preview_code(csv_filename),
                    "language": "python",
                    "clearContext": False,
                })
                preview_stdout = ""
                for event in preview_resp.get("stream", []):
                    result = event.get("result", {})
                    if result.get("isError", False):
                        continue
                    preview_stdout += result.get("structuredContent", {}).get("stdout", "")
                schema_preview, _ = _extract_schema_preview(preview_stdout)
            except Exception as e:
                logger.warning(f"Schema preview failed for {csv_filename}: {e}")

            # 6. Execute user code
            response = code_interpreter.invoke("executeCode", {
                "code": python_code,
                "language": "python",
                "clearContext": False,
            })

            execution_output = ""
            for event in response.get("stream", []):
                result = event.get("result", {})
                if result.get("isError", False):
                    error_msg = _clean_stderr(result.get("structuredContent", {}).get("stderr", ""))
                    error_text = f"❌ Code execution failed:\n```\n{error_msg}\n```"

                    # Targeted hint for the most common wrong-filename error:
                    # the model wrote `pd.read_csv('FY_27_Ledger.xlsx', ...)`
                    # but in the sandbox the file lives as `FY_27_Ledger.csv`
                    # (see docstring: XLSX sources are pre-converted). Naming
                    # this out explicitly is much more effective than relying
                    # on the model to infer it from the schema footer.
                    if (
                        is_xlsx
                        and "FileNotFoundError" in error_msg
                        and filename in error_msg
                    ):
                        error_text += (
                            f"\n\n**Hint:** In the sandbox, the XLSX source "
                            f"`{filename}` is loaded as `{csv_filename}`. "
                            f"Retry with `pd.read_csv('{csv_filename}', "
                            f"low_memory=False)`."
                        )

                    if schema_preview:
                        # Drop the first_row dump on errors — the load line +
                        # column list is enough for the retry, first_row is
                        # ~1K tokens of bloat on a path that's already costing
                        # a round-trip.
                        trimmed_schema = _strip_first_row(schema_preview)
                        error_text += f"\n\nDataset info (use the `load:` line verbatim):\n```\n{trimmed_schema}\n```"
                    else:
                        error_text += f"\n\nTry: `pd.read_csv('{csv_filename}', low_memory=False)`"
                    if multi_sheet_note:
                        error_text += f"\n\n{multi_sheet_note}"
                    return {"content": [{"text": error_text}], "status": "error"}
                stdout = result.get("structuredContent", {}).get("stdout", "")
                if stdout:
                    execution_output += stdout

            # 7. Download chart if requested
            success_text = _truncate_output(execution_output) or "✅ Code executed successfully (no output)."
            if schema_preview:
                success_text = f"{success_text}\n\n---\nDataset: {schema_preview.splitlines()[0] if schema_preview else ''}"
            if multi_sheet_note:
                success_text = f"{success_text}\n{multi_sheet_note}"

            if output_filename and output_filename.endswith(".png"):
                try:
                    dl_response = code_interpreter.invoke("readFiles", {"paths": [output_filename]})
                    file_content = None
                    for event in dl_response.get("stream", []):
                        result = event.get("result", {})
                        if "content" in result:
                            for block in result["content"]:
                                if "data" in block:
                                    file_content = block["data"]
                                elif "resource" in block and "blob" in block["resource"]:
                                    file_content = block["resource"]["blob"]
                                if file_content:
                                    break
                        if file_content:
                            break

                    if file_content:
                        return {
                            "content": [
                                {"text": success_text},
                                {"image": {"format": "png", "source": {"bytes": file_content}}},
                            ],
                            "status": "success",
                        }
                except Exception as e:
                    logger.warning(f"Failed to download chart {output_filename}: {e}")

            return {
                "content": [{"text": success_text}],
                "status": "success",
            }

        finally:
            try:
                code_interpreter.stop()
            except Exception:
                pass

    return analyze_spreadsheet


async def _find_file(filename: str, assistant_id: Optional[str], session_id: str) -> Optional[Dict[str, Any]]:
    """Find a file by name in accessible sources. Returns file info or None.

    Matches are tolerant to XLSX ↔ CSV aliasing: if the model asks for
    ``foo.csv`` but only ``foo.xlsx`` exists (because the sandbox converts
    XLSX → CSV and the model copied the sandbox name into the ``filename``
    param on retry), we treat them as the same file. Prevents the common
    round-trip loop where analyze_spreadsheet rejects a reasonable guess
    and forces the model to call list_spreadsheets (#206).
    """
    candidates: list[Dict[str, Any]] = []
    if assistant_id:
        candidates.extend(await _get_kb_files(assistant_id))
    candidates.extend(await _get_session_files(session_id))

    target_lower = filename.lower()
    target_stem, _ = os.path.splitext(target_lower)

    # First pass: exact match (case-insensitive).
    for f in candidates:
        if f["filename"].lower() == target_lower:
            return f

    # Second pass: same stem, tabular extension. Covers foo.csv -> foo.xlsx
    # and foo.xlsx -> foo.csv. Only applies to tabular files so we don't
    # accidentally alias foo.pdf to foo.docx.
    from apis.shared.files.models import is_tabular_file

    if target_stem and any(target_lower.endswith(ext) for ext in (".csv", ".xls", ".xlsx")):
        for f in candidates:
            cand_lower = f["filename"].lower()
            cand_stem, _ = os.path.splitext(cand_lower)
            if cand_stem == target_stem and is_tabular_file(f["filename"], f.get("content_type", "")):
                return f

    return None


def _download_file(file_info: Dict[str, Any]) -> bytes:
    """Download file bytes from S3."""
    region = os.environ.get("AWS_REGION", "us-west-2")
    s3 = boto3.client("s3", region_name=region)

    if file_info["source"] == "knowledge_base":
        bucket = os.environ.get("S3_ASSISTANTS_DOCUMENTS_BUCKET_NAME")
        if not bucket:
            raise ValueError("S3_ASSISTANTS_DOCUMENTS_BUCKET_NAME not configured")
    else:
        bucket = file_info.get("s3_bucket")
        if not bucket:
            raise ValueError("S3 bucket not found in file metadata")

    response = s3.get_object(Bucket=bucket, Key=file_info["s3_key"])
    return response["Body"].read()
